import pandas as pd
import os
from dotenv import find_dotenv, load_dotenv
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from azure.storage.blob import BlobServiceClient


load_dotenv(find_dotenv())
SAS_TOKEN = os.environ["SAS_TOKEN"]

account_name = "pythonwhatsappbotstorage"
container_name = "data"
blob_name_load = "df_sales.xlsx"
sas_token = SAS_TOKEN
file_path_upload = "income_statement_visual.xlsx"
blob_name_upload = "income_statement_visual.xlsx"


def load_data_from_azure(account_name, container_name, blob_name, sas_token):
    try:
        # Create a BlobServiceClient using the SAS Token and get a client to interact with the specified blob
        blob_service_client = BlobServiceClient(account_url=f"https://{account_name}.blob.core.windows.net", credential=sas_token)
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
        
        # Download the blob to a local memory stream
        with open(blob_name_load, "wb") as download_file:
            download_file.write(blob_client.download_blob().readall())
        df = pd.read_excel(blob_name_load)

        return df

    except Exception as e:
        # Handle generic exceptions which could be related to Azure access issues, network problems, etc.
        print(f"Failed to load data from Azure Blob Storage: {e}")
        return None


def create_income_statement(df):
    try:
        def load_and_process_data(df):
            try:
                data = df
            except FileNotFoundError:
                return None, "File not found. Please check the file path."
            except Exception as e:
                return None, f"Error reading the Excel file: {e}"

            try:
                data['Date'] = pd.to_datetime(data['Date'])
                data['Year'] = data['Date'].dt.year
                data['Month'] = data['Date'].dt.month.map({
                    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
                    7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
                })
                data['Total_Sales'] = data['Sales_Quantity'] * data['unit_sales_price']
                data['Total_Cost'] = data['Sales_Quantity'] * data['unit_purchase_cost']
                data['Profit'] = data['Total_Sales'] - data['Total_Cost']
            except KeyError as e:
                return None, f"Missing expected column in data: {e}"
            except Exception as e:
                return None, f"Error processing data: {e}"

            return data, None

        def aggregate_monthly(data):
            try:
                aggregates = data.groupby(['Month']).agg({
                    'Total_Sales': 'sum',
                    'Total_Cost': 'sum',
                    'Profit': 'sum'
                }).rename(columns={
                    'Total_Sales': 'Ventas',
                    'Total_Cost': 'Costo',
                    'Profit': 'Utilidad'
                })
                aggregates['Margen de Utilidad (%)'] = (aggregates['Utilidad'] / aggregates['Ventas'])
                months_order = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
                aggregates = aggregates.reindex(months_order, fill_value=0)
            except Exception as e:
                return None, f"Error aggregating monthly data: {e}"

            return aggregates, None

        def format_income_statement(aggregate_data):
            try:
                aggregate_data.loc['Total_Año'] = aggregate_data.sum(numeric_only=True)
                aggregate_data.loc['Total_Año', 'Margen de Utilidad (%)'] = aggregate_data.loc['Total_Año', 'Utilidad'] / aggregate_data.loc['Total_Año', 'Ventas']
                formatted = aggregate_data.T
            except Exception as e:
                return None, f"Error formatting income statement: {e}"

            return formatted, None

        def calculate_growth(all_years_data):
            try:
                for i in range(len(all_years_data) - 1):
                    current_year, current_statement = all_years_data[i]
                    prev_year, prev_statement = all_years_data[i + 1]
                    for month in current_statement.columns:
                        if month != 'Total_Año':
                            prev_year_sales = prev_statement.loc['Ventas', month]
                            current_year_sales = current_statement.loc['Ventas', month]
                            sales_growth = (current_year_sales - prev_year_sales) / prev_year_sales if prev_year_sales != 0 else None
                            if sales_growth == -1:
                                sales_growth = None
                            current_statement.loc['Crecimiento de Ventas (%)', month] = sales_growth

                            prev_year_profit = prev_statement.loc['Utilidad', month]
                            current_year_profit = current_statement.loc['Utilidad', month]
                            profit_growth = (current_year_profit - prev_year_profit) / prev_year_profit if prev_year_profit != 0 else None
                            if profit_growth == -1:
                                profit_growth = None
                            current_statement.loc['Crecimiento de Utilidad (%)', month] = profit_growth

                    prev_year_total_sales = prev_statement.loc['Ventas', 'Total_Año']
                    current_year_total_sales = current_statement.loc['Ventas', 'Total_Año']
                    yearly_sales_growth = (current_year_total_sales - prev_year_total_sales) / prev_year_total_sales if prev_year_total_sales != 0 else None
                    if yearly_sales_growth == -1:
                        yearly_sales_growth = None
                    current_statement.loc['Crecimiento de Ventas (%)', 'Total_Año'] = yearly_sales_growth

                    prev_year_total_profit = prev_statement.loc['Utilidad', 'Total_Año']
                    current_year_total_profit = current_statement.loc['Utilidad', 'Total_Año']
                    yearly_profit_growth = (current_year_total_profit - prev_year_total_profit) / prev_year_total_profit if prev_year_total_profit != 0 else None
                    if yearly_profit_growth == -1:
                        yearly_profit_growth = None
                    current_statement.loc['Crecimiento de Utilidad (%)', 'Total_Año'] = yearly_profit_growth
            except Exception as e:
                return None, f"Error calculating growth: {e}"

            return all_years_data, None

        def generate_income_statements(data):
            try:
                years = sorted(data['Year'].unique(), reverse=True)  # Sort years in descending order
                all_years_data = []
                for year in years:
                    yearly_data = data[data['Year'] == year]
                    monthly_data, error = aggregate_monthly(yearly_data)
                    if error:
                        return None, error
                    income_statement, error = format_income_statement(monthly_data)
                    if error:
                        return None, error
                    all_years_data.append((year, income_statement))
                
                # Calculate Sales and Profit Growth
                all_years_data, error = calculate_growth(all_years_data)
                if error:
                    return None, error

            except Exception as e:
                return None, f"Error generating income statements: {e}"

            return all_years_data, None

        # Process the data
        data, error = load_and_process_data(df)
        if error:
            return error

        all_years_data, error = generate_income_statements(data)
        if error:
            return error

        # Determine the date range for the title
        min_year = data['Date'].min().strftime('%Y')
        max_year = data['Date'].max().strftime('%Y')

        # Define styles
        no_border = Border(left=Side(border_style=None),
                           right=Side(border_style=None),
                           top=Side(border_style=None),
                           bottom=Side(border_style=None))

        bottom_border = Border(bottom=Side(border_style='thin'))

        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        font_color = '156082'
        font_name = 'Aptos Narrow'  # Change this to your desired font type

        try:
            # Save to Excel with each year stacked vertically and formatting
            with pd.ExcelWriter('income_statement_visual.xlsx', engine='openpyxl') as writer:
                # Create the title
                title = f'Estado de Resultados mensual {min_year} - {max_year}'
                
                row_start = 4  # Start after the title
                for year, income_statement in all_years_data:
                    # Add the Empty row before Sales Growth (%)
                    empty_row = pd.DataFrame(index=[''], columns=income_statement.columns).astype(income_statement.dtypes)
                    income_statement = pd.concat([income_statement.iloc[:4], empty_row, income_statement.iloc[4:]])
                    income_statement.to_excel(writer, startrow=row_start)
                    ws = writer.sheets['Sheet1']
                    
                    # Apply white fill to all cells
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 10, min_col=1, max_col=ws.max_column + 2):
                        for cell in row:
                            cell.fill = white_fill

                    # Add the title to the top
                    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(income_statement.columns) + 1)
                    title_cell = ws.cell(row=2, column=2)
                    title_cell.value = title
                    title_cell.font = Font(name=font_name, size=22, bold=True)  # Set font type
                    title_cell.alignment = Alignment(horizontal='left')

                    # Format the year title
                    year_title_cell = f'B{row_start}'
                    ws[year_title_cell] = year  # Set the year in the first cell
                    ws[year_title_cell].font = Font(name=font_name, bold=True, size=20)  # Make the year title bold and larger
                    ws[year_title_cell].border = no_border  # Remove border for the year title

                    # Set column widths
                    ws.column_dimensions['A'].width = 25  # Set the width for the first column
                    for col in range(2, len(income_statement.columns) + 2):
                        col_letter = ws.cell(row=row_start + 1, column=col).column_letter
                        ws.column_dimensions[col_letter].width = 11

                    # Center the month headers and add bottom border
                    for col in range(2, len(income_statement.columns) + 2):
                        cell = ws.cell(row=row_start + 1, column=col)
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(name=font_name, bold=True)  # Set font type
                        cell.number_format = '@'  # Set text format
                        cell.border = bottom_border  # Add bottom border for month headers

                    # Add header for Total_Año (skip this as it's causing an extra title)
                    cell = ws.cell(row=row_start + 1, column=len(income_statement.columns) + 1)
                    cell.value = 'Total_Año'
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(name=font_name, bold=True)  # Set font type
                    cell.number_format = '@'  # Set text format
                    cell.border = bottom_border  # Add bottom border

                    # Apply different font styles to row titles
                    row_titles = {
                        'Ventas': (row_start + 2, False),
                        'Costo': (row_start + 3, False),
                        'Utilidad': (row_start + 4, True),
                        'Margen de Utilidad (%)': (row_start + 5, True),
                        '': (row_start + 6, False),
                        'Crecimiento de Ventas (%)': (row_start + 7, False),
                        'Crecimiento de Utilidad (%)': (row_start + 8, False),
                    }
                    for row_name, (row_number, is_bold) in row_titles.items():
                        cell = ws.cell(row=row_number, column=1)
                        cell.font = Font(name=font_name, bold=is_bold, color=font_color if row_name in ['Crecimiento de Ventas (%)', 'Crecimiento de Utilidad (%)'] else None)
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = no_border  # Remove border for row titles

                    # Apply formatting to all cells in Profit, Margin (%), Sales Growth (%), and Profit Growth (%) rows
                    growth_rows = [row_start + 7, row_start + 8]  # Rows for Sales Growth (%) and Profit Growth (%)
                    for row in growth_rows:
                        for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                            cell = ws.cell(row=row, column=col)
                            cell.font = Font(name=font_name, color=font_color)

                    bold_rows = [row_start + 4, row_start + 5]  # Rows for Profit and Margin (%)
                    for row in bold_rows:
                        for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                            cell = ws.cell(row=row, column=col)
                            cell.font = Font(name=font_name, bold=True, italic=(row == row_start + 5))

                    # Apply formatting to all other cells
                    for row in ws.iter_rows(min_row=row_start + 2, max_row=row_start + len(income_statement.index) + 1,
                                            min_col=2, max_col=len(income_statement.columns) + 1):  # Include Total_Año column
                        for cell in row:
                            if cell.value == 0:
                                cell.value = '-'
                                cell.alignment = Alignment(horizontal='center')
                                cell.number_format = '@'  # Set text format
                                cell.border = no_border  # Remove border for '-' cells
                            else:
                                cell.alignment = Alignment(horizontal='right')
                                cell.font = Font(name=font_name)  # Set font type
                                cell.border = no_border  # Remove border for all other cells

                    # Add bottom border to the entire Cost row
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=row_start + 3, column=col)
                        cell.border = bottom_border  # Add bottom border for Cost row

                    # Apply number format to Sales, Cost, and Profit
                    format_rows = {
                        'Ventas': row_start + 2,
                        'Costo': row_start + 3,
                        'Utilidad': row_start + 4,
                    }
                    for row_name, row_number in format_rows.items():
                        for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                            cell = ws.cell(row=row_number, column=col)
                            if cell.value != '-':
                                cell.number_format = '#,##0'  # Format numbers with comma for thousands
                                cell.font = Font(name=font_name)  # Set font type

                    # Format the profit margin cells
                    margin_row = row_start + 5  # Margin row
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=margin_row, column=col)
                        if cell.value == '-':
                            cell.alignment = Alignment(horizontal='center')
                            cell.number_format = '@'  # Set text format
                        else:
                            cell.number_format = '0.00%'  # Set percentage format
                            cell.font = Font(name=font_name, bold=True, italic=True)  # Set bold and italic format

                    # Format the empty row cells
                    empty_row = row_start + 6  # Empty row
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=empty_row, column=col)
                        cell.value = ''
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(name=font_name)  # Set font type
                        cell.number_format = '@'  # Set text format

                    # Format the sales growth cells
                    sales_growth_row = row_start + 7  # Sales Growth row
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=sales_growth_row, column=col)
                        if cell.value is None:
                            cell.value = ''
                        elif cell.value == '-':
                            cell.alignment = Alignment(horizontal='center')
                            cell.number_format = '@'  # Set text format
                        else:
                            cell.number_format = '0.00%'  # Set percentage format
                            cell.font = Font(name=font_name, color=font_color)

                    # Format the profit growth cells
                    profit_growth_row = row_start + 8  # Profit Growth row
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=profit_growth_row, column=col)
                        if cell.value is None:
                            cell.value = ''
                        elif cell.value == '-':
                            cell.alignment = Alignment(horizontal='center')
                            cell.number_format = '@'  # Set text format
                        else:
                            cell.number_format = '0.00%'
                            cell.font = Font(name=font_name, color=font_color)

                    profit_row = row_start + 4
                    for col in range(2, len(income_statement.columns) + 2):  # Include Total_Año column
                        cell = ws.cell(row=profit_row, column=col)
                        cell.font = Font(name=font_name, bold=True)

                    # Increment the row start for the next year's data
                    row_start += len(income_statement.index) + 4  # Add blank rows for spacing

        except Exception as e:
            return f"Error saving to Excel: {e}"

        return "Income statement created successfully and saved as 'income_statement_visual.xlsx'."

    except Exception as e:
        return f"An unexpected error occurred: {e}"
    

def upload_to_azure_with_sas(file_path, account_name, container_name, blob_name, sas_token):
    try:
        # Create the BlobServiceClient object using the SAS token and the account URL
        account_url = f"https://{account_name}.blob.core.windows.net"
        blob_service_client = BlobServiceClient(account_url=account_url, credential=sas_token)

        # Create a blob client using the local file name as the name for the blob
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

        print(f"Uploading to Azure Storage as blob:\n\t{blob_name}")

        # Upload the created file
        with open(file_path, "rb") as data:
            blob_client.upload_blob(data, overwrite=True)

        return f"File {file_path} uploaded to Azure Blob Storage successfully as {blob_name}."
    except Exception as ex:
        return f"An error occurred: {ex}"


df_sales = load_data_from_azure(account_name, container_name, blob_name_load, sas_token)
create_income_statement(df_sales)
message = upload_to_azure_with_sas(file_path_upload, account_name, container_name, blob_name_upload, sas_token)
print(message)
